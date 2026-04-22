"""
Quiz Router — Full CRUD + Attempt tracking + Teacher stats.
Prefix: /api/quizzes
"""
import logging
import json
import re
import time as _time
from datetime import datetime, timezone
from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

# #region agent log
def _qlog(msg, data, hyp="H2"):
    try:
        with open("/workspace/debug-1b57e8.log", "a") as _f:
            _f.write(json.dumps({"sessionId":"1b57e8","timestamp":int(_time.time()*1000),"location":"quiz.py","message":msg,"data":data,"hypothesisId":hyp}) + "\n")
    except Exception: pass
# #endregion

from app.database.db import get_db
from app.dependencies.deps import get_current_user, check_enrollment
from app.models.user import User, UserRole
from app.models.subject import Enrollment, EnrollmentRole, Course
from app.models.schedule import GeneratedQuiz, QuizAttempt, ExamSchedule, ExamType
from app.services.llm_service import LLMService
from app.services.vector_service import get_vector_service
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/quizzes", tags=["Quizzes"])


# ─── Pydantic Schemas ────────────────────────────────────

class QuizQuestionSchema(BaseModel):
    question: str
    options: dict        # {"A": "...", "B": "...", "C": "...", "D": "..."}
    correct_answer: str  # "A" | "B" | "C" | "D"
    explanation: str = ""


class QuizCreateManual(BaseModel):
    course_id: int
    title: str
    questions: List[QuizQuestionSchema]


class QuizCreateAI(BaseModel):
    course_id: int
    title: str
    topic: str = ""
    num_questions: int = 5


class AttemptSubmit(BaseModel):
    answers: dict                       # {"0": "A", "1": "C", ...}
    time_spent_seconds: Optional[int] = None
    started_at: Optional[str] = None


class QuizResponse(BaseModel):
    id: int
    course_id: int
    title: str
    question_count: int
    is_ai_generated: bool
    created_by: Optional[str] = None
    created_at: str

    class Config:
        from_attributes = True


class QuizDetailResponse(QuizResponse):
    questions: List[dict]


class AttemptResponse(BaseModel):
    id: int
    quiz_id: int
    user_id: int
    user_name: Optional[str] = None
    score: int
    total_questions: int
    time_spent_seconds: Optional[int]
    completed_at: Optional[str]
    created_at: str

    class Config:
        from_attributes = True


class CourseStatsResponse(BaseModel):
    course_id: int
    total_quizzes: int
    total_attempts: int
    average_score_pct: float
    quizzes: List[dict]


class ExamScheduleUpsert(BaseModel):
    exam_type: ExamType
    exam_date: datetime
    duration_minutes: int = Field(gt=0, le=600)
    location: Optional[str] = None
    notes: Optional[str] = None


# ─── Helpers ────────────────────────────────────────────

def _quiz_to_response(quiz: GeneratedQuiz) -> dict:
    return {
        "id": quiz.id,
        "course_id": quiz.course_id,
        "title": quiz.title,
        "question_count": len(quiz.questions) if quiz.questions else 0,
        "is_ai_generated": bool(quiz.is_ai_generated),
        "created_by": quiz.user.full_name if quiz.user else None,
        "created_at": quiz.created_at.isoformat() if quiz.created_at else None,
    }


def _attempt_to_response(attempt: QuizAttempt) -> dict:
    return {
        "id": attempt.id,
        "quiz_id": attempt.quiz_id,
        "user_id": attempt.user_id,
        "user_name": attempt.user.full_name if attempt.user else None,
        "score": attempt.score,
        "total_questions": attempt.total_questions,
        "time_spent_seconds": attempt.time_spent_seconds,
        "completed_at": attempt.completed_at.isoformat() if attempt.completed_at else None,
        "created_at": attempt.created_at.isoformat() if attempt.created_at else None,
    }


def _schedule_to_response(schedule: ExamSchedule) -> dict:
    return {
        "id": schedule.id,
        "course_id": schedule.course_id,
        "exam_type": schedule.exam_type.value if schedule.exam_type else None,
        "exam_date": schedule.exam_date.isoformat() if schedule.exam_date else None,
        "duration_minutes": schedule.duration_minutes,
        "location": schedule.location,
        "notes": schedule.notes,
    }


def _require_course_access(db: Session, user: User, course_id: int):
    """Raise 403 if user has no access to course."""
    if not check_enrollment(db, user, course_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bạn không có quyền truy cập môn học này")


def _require_teacher_or_admin(user: User):
    if user.role == UserRole.student:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Chỉ giáo viên hoặc admin mới có thể xuất báo cáo này",
        )


def _require_admin(user: User):
    if user.role != UserRole.admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Chỉ admin mới có thể quản lý lịch thi",
        )


def _normalize_schedule_datetime(value: datetime) -> datetime:
    if value.tzinfo is not None:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value


def _safe_filename_segment(s: str, max_len: int = 80) -> str:
    s = re.sub(r'[<>:"/\\|?*\n\r\t]', "_", (s or "").strip())
    return (s[:max_len] if s else "export")


def _format_time_spent(sec: Optional[int]) -> str:
    if sec is None:
        return "—"
    m, s = divmod(int(sec), 60)
    return f"{m}p {s}s"


def _excel_response(content: bytes, filename_base: str) -> Response:
    safe = _safe_filename_segment(filename_base, 120)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe}.xlsx"',
        },
    )


def _build_quiz_attempts_workbook(
    quiz_title: str,
    course_name: str,
    attempts: List[QuizAttempt],
) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Kết quả"

    now_str = datetime.utcnow().strftime("%d/%m/%Y %H:%M")
    ws["A1"] = f"Kết quả Quiz: {quiz_title}"
    ws["A2"] = f"Môn học: {course_name}    |    Xuất ngày: {now_str}"
    ws["A4"] = "STT"
    ws["B4"] = "Họ và tên"
    ws["C4"] = "Điểm"
    ws["D4"] = "Tổng câu"
    ws["E4"] = "Phần trăm (%)"
    ws["F4"] = "Thời gian làm bài"
    ws["G4"] = "Ngày nộp bài"

    row = 5
    pcts: List[float] = []
    pass_count = 0
    for idx, a in enumerate(attempts, start=1):
        total = a.total_questions or 0
        pct = (a.score / total * 100.0) if total > 0 else 0.0
        pcts.append(pct)
        if total > 0 and (a.score / total) >= 0.5:
            pass_count += 1
        done = a.completed_at or a.created_at
        done_str = done.strftime("%d/%m/%Y %H:%M") if done else ""
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=a.user.full_name if a.user else f"User #{a.user_id}")
        ws.cell(row=row, column=3, value=a.score)
        ws.cell(row=row, column=4, value=total)
        ws.cell(row=row, column=5, value=round(pct, 1))
        ws.cell(row=row, column=6, value=_format_time_spent(a.time_spent_seconds))
        ws.cell(row=row, column=7, value=done_str)
        row += 1

    summary_row = row + 1
    n = len(attempts)
    avg_pct = round(sum(pcts) / n, 1) if n else 0.0
    pass_rate = round(pass_count / n * 100, 1) if n else 0.0
    ws.cell(row=summary_row, column=1, value="Tổng kết")
    ws.cell(row=summary_row, column=2, value=f"Điểm TB (%): {avg_pct}")
    ws.cell(row=summary_row, column=5, value=f"Tỷ lệ đạt (≥50%): {pass_rate}%")

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_course_stats_workbook(
    course_name: str,
    total_quizzes: int,
    total_attempts: int,
    average_score_pct: float,
    quiz_rows: List[dict],
) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Tổng quan"
    ws1["A1"] = "Tên môn học"
    ws1["B1"] = course_name
    ws1["A2"] = "Tổng số quiz"
    ws1["B2"] = total_quizzes
    ws1["A3"] = "Tổng lượt làm bài"
    ws1["B3"] = total_attempts
    ws1["A4"] = "Điểm trung bình toàn khóa (%)"
    ws1["B4"] = round(average_score_pct, 1)

    ws2 = wb.create_sheet("Chi tiết từng Quiz")
    ws2["A1"] = "Tên quiz"
    ws2["B1"] = "Số câu"
    ws2["C1"] = "Lượt làm"
    ws2["D1"] = "Điểm TB (%)"
    r = 2
    for q in quiz_rows:
        ws2.cell(row=r, column=1, value=q.get("title", ""))
        ws2.cell(row=r, column=2, value=q.get("question_count", 0))
        ws2.cell(row=r, column=3, value=q.get("attempt_count", 0))
        ws2.cell(row=r, column=4, value=q.get("average_score_pct", 0))
        r += 1

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_my_attempts_workbook(rows: List[dict]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Lịch sử làm bài"

    ws["A1"] = "STT"
    ws["B1"] = "Tên Quiz"
    ws["C1"] = "Điểm"
    ws["D1"] = "Tổng câu"
    ws["E1"] = "Phần trăm (%)"
    ws["F1"] = "Thời gian"
    ws["G1"] = "Ngày hoàn thành"

    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=row.get("quiz_title", ""))
        ws.cell(row=i, column=3, value=row.get("score"))
        ws.cell(row=i, column=4, value=row.get("total_questions"))
        ws.cell(row=i, column=5, value=row.get("score_pct"))
        ws.cell(row=i, column=6, value=row.get("time_str"))
        ws.cell(row=i, column=7, value=row.get("done_str"))

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _notify_students_new_quiz(db: Session, quiz: GeneratedQuiz):
    """Fire-and-forget Celery task notifying enrolled students of a new quiz."""
    try:
        from app.celery_app import celery_app
        from app.models.subject import Enrollment, EnrollmentRole
        enrolled_ids = [
            e.user_id for e in db.query(Enrollment).filter(
                Enrollment.course_id == quiz.course_id,
                Enrollment.role == EnrollmentRole.student,
            ).all()
        ]
        if enrolled_ids:
            course_name = quiz.course.name if quiz.course else f"Môn #{quiz.course_id}"
            celery_app.send_task(
                'create_notifications',
                kwargs=dict(
                    user_ids=enrolled_ids,
                    type="quiz",
                    title="Quiz mới cho môn học",
                    message=f"Quiz '{quiz.title}' vừa được tạo trong môn {course_name}",
                    related_type="quiz",
                    related_id=quiz.id,
                ),
            )
    except Exception as _ne:
        logger.warning(f"Failed to queue quiz notification: {_ne}")


# ─── Endpoints ──────────────────────────────────────────

@router.get("/me/attempts")
def get_my_attempts(
    course_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get all quiz attempts for the current user."""
    query = (
        db.query(QuizAttempt)
        .options(joinedload(QuizAttempt.user))
        .filter(QuizAttempt.user_id == current_user.id)
    )
    if course_id:
        query = query.join(GeneratedQuiz).filter(GeneratedQuiz.course_id == course_id)
    attempts = query.order_by(QuizAttempt.created_at.desc()).all()
    return {"items": [_attempt_to_response(a) for a in attempts], "total": len(attempts)}


@router.get("/me/attempts/export")
def export_my_attempts_excel(
    course_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Xuất lịch sử làm quiz của người dùng hiện tại (file Excel)."""
    query = (
        db.query(QuizAttempt)
        .options(joinedload(QuizAttempt.user), joinedload(QuizAttempt.quiz))
        .filter(QuizAttempt.user_id == current_user.id)
    )
    if course_id is not None:
        query = query.join(GeneratedQuiz).filter(GeneratedQuiz.course_id == course_id)
        _require_course_access(db, current_user, course_id)

    attempts = query.order_by(QuizAttempt.created_at.desc()).all()

    rows = []
    for a in attempts:
        quiz = a.quiz
        quiz_title = quiz.title if quiz else f"Quiz #{a.quiz_id}"
        total = a.total_questions or 0
        pct = round((a.score / total * 100), 1) if total > 0 else 0.0
        done = a.completed_at or a.created_at
        done_str = done.strftime("%d/%m/%Y %H:%M") if done else ""
        rows.append({
            "quiz_title": quiz_title,
            "score": a.score,
            "total_questions": total,
            "score_pct": pct,
            "time_str": _format_time_spent(a.time_spent_seconds),
            "done_str": done_str,
        })

    content = _build_my_attempts_workbook(rows)
    fname = f"my_quiz_history_{datetime.utcnow().strftime('%Y%m%d')}"
    return _excel_response(content, fname)


@router.get("/schedules/course/{course_id}")
def get_exam_schedules(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get exam schedules for a course."""
    _require_course_access(db, current_user, course_id)
    
    schedules = (
        db.query(ExamSchedule)
        .filter(ExamSchedule.course_id == course_id)
        .order_by(ExamSchedule.exam_date)
        .all()
    )
    return {
        "items": [_schedule_to_response(s) for s in schedules],
        "total": len(schedules),
    }


@router.post("/schedules/course/{course_id}", status_code=status.HTTP_201_CREATED)
def create_exam_schedule(
    course_id: int,
    payload: ExamScheduleUpsert,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create an exam schedule. Admin only."""
    _require_admin(current_user)

    course = db.query(Course).filter(Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn học")

    schedule = ExamSchedule(
        course_id=course_id,
        exam_type=payload.exam_type,
        exam_date=_normalize_schedule_datetime(payload.exam_date),
        duration_minutes=payload.duration_minutes,
        location=payload.location.strip() if payload.location and payload.location.strip() else None,
        notes=payload.notes.strip() if payload.notes and payload.notes.strip() else None,
    )
    db.add(schedule)
    db.commit()
    db.refresh(schedule)
    return _schedule_to_response(schedule)


@router.put("/schedules/{schedule_id}")
def update_exam_schedule(
    schedule_id: int,
    payload: ExamScheduleUpsert,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update an exam schedule. Admin only."""
    _require_admin(current_user)

    schedule = db.query(ExamSchedule).filter(ExamSchedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Không tìm thấy lịch thi")

    schedule.exam_type = payload.exam_type
    schedule.exam_date = _normalize_schedule_datetime(payload.exam_date)
    schedule.duration_minutes = payload.duration_minutes
    schedule.location = payload.location.strip() if payload.location and payload.location.strip() else None
    schedule.notes = payload.notes.strip() if payload.notes and payload.notes.strip() else None
    db.commit()
    db.refresh(schedule)
    return _schedule_to_response(schedule)


@router.delete("/schedules/{schedule_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_exam_schedule(
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete an exam schedule. Admin only."""
    _require_admin(current_user)

    schedule = db.query(ExamSchedule).filter(ExamSchedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Không tìm thấy lịch thi")

    db.delete(schedule)
    db.commit()


@router.get("/stats/course/{course_id}")
def get_course_stats(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get quiz statistics for a course. Accessible to all enrolled users."""
    _require_course_access(db, current_user, course_id)

    quizzes = db.query(GeneratedQuiz).filter(GeneratedQuiz.course_id == course_id).all()
    quiz_ids = [q.id for q in quizzes]

    total_attempts = db.query(QuizAttempt).filter(QuizAttempt.quiz_id.in_(quiz_ids)).count() if quiz_ids else 0

    avg_score_pct = 0.0
    if quiz_ids and total_attempts > 0:
        attempts = db.query(QuizAttempt).filter(QuizAttempt.quiz_id.in_(quiz_ids)).all()
        if attempts:
            avg_score_pct = sum(
                (a.score / a.total_questions * 100) for a in attempts if a.total_questions > 0
            ) / len(attempts)

    quizzes_stats = []
    for q in quizzes:
        q_attempts = db.query(QuizAttempt).filter(QuizAttempt.quiz_id == q.id).all()
        q_avg = 0.0
        if q_attempts:
            q_avg = sum((a.score / a.total_questions * 100) for a in q_attempts if a.total_questions > 0) / len(q_attempts)
        quizzes_stats.append({
            "id": q.id,
            "title": q.title,
            "question_count": len(q.questions) if q.questions else 0,
            "attempt_count": len(q_attempts),
            "average_score_pct": round(q_avg, 1),
        })

    return {
        "course_id": course_id,
        "total_quizzes": len(quizzes),
        "total_attempts": total_attempts,
        "average_score_pct": round(avg_score_pct, 1),
        "quizzes": quizzes_stats,
    }


@router.get("/stats/course/{course_id}/export")
def export_course_stats_excel(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Xuất Excel tổng hợp quiz theo môn học (giáo viên / admin)."""
    _require_teacher_or_admin(current_user)
    _require_course_access(db, current_user, course_id)

    course = db.query(Course).filter(Course.id == course_id).first()
    course_name = course.name if course else f"Môn #{course_id}"

    quizzes = db.query(GeneratedQuiz).filter(GeneratedQuiz.course_id == course_id).all()
    quiz_ids = [q.id for q in quizzes]

    total_attempts = db.query(QuizAttempt).filter(QuizAttempt.quiz_id.in_(quiz_ids)).count() if quiz_ids else 0

    avg_score_pct = 0.0
    if quiz_ids and total_attempts > 0:
        attempts_all = db.query(QuizAttempt).filter(QuizAttempt.quiz_id.in_(quiz_ids)).all()
        if attempts_all:
            avg_score_pct = sum(
                (a.score / a.total_questions * 100) for a in attempts_all if a.total_questions > 0
            ) / len(attempts_all)

    quiz_rows = []
    for q in quizzes:
        q_attempts = db.query(QuizAttempt).filter(QuizAttempt.quiz_id == q.id).all()
        q_avg = 0.0
        if q_attempts:
            q_avg = sum(
                (a.score / a.total_questions * 100) for a in q_attempts if a.total_questions > 0
            ) / len(q_attempts)
        quiz_rows.append({
            "title": q.title,
            "question_count": len(q.questions) if q.questions else 0,
            "attempt_count": len(q_attempts),
            "average_score_pct": round(q_avg, 1),
        })

    content = _build_course_stats_workbook(
        course_name=course_name,
        total_quizzes=len(quizzes),
        total_attempts=total_attempts,
        average_score_pct=round(avg_score_pct, 1),
        quiz_rows=quiz_rows,
    )
    fname = f"quiz_stats_{_safe_filename_segment(course_name, 40)}_{datetime.utcnow().strftime('%Y%m%d')}"
    return _excel_response(content, fname)


@router.get("/")
def list_quizzes(
    course_id: Optional[int] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    List quizzes.
    - Admin: sees all quizzes (optionally filtered by course)
    - Teacher/Student: sees quizzes for courses they're enrolled in
    """
    if current_user.role == UserRole.admin:
        query = db.query(GeneratedQuiz).options(joinedload(GeneratedQuiz.user))
        if course_id:
            query = query.filter(GeneratedQuiz.course_id == course_id)
    else:
        enrolled_course_ids = [
            e.course_id for e in db.query(Enrollment).filter(Enrollment.user_id == current_user.id).all()
        ]
        query = db.query(GeneratedQuiz).options(joinedload(GeneratedQuiz.user)).filter(GeneratedQuiz.course_id.in_(enrolled_course_ids))
        if course_id:
            if course_id not in enrolled_course_ids:
                raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập môn học này")
            query = query.filter(GeneratedQuiz.course_id == course_id)

    try:
        total = query.count()
        quizzes = query.order_by(GeneratedQuiz.created_at.desc()).offset(skip).limit(limit).all()
        # #region agent log
        _qlog("list_quizzes_ok", {"total": total, "course_id": course_id, "user_role": current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role)}, "H2")
        # #endregion
        return {"items": [_quiz_to_response(q) for q in quizzes], "total": total}
    except Exception as e:
        # #region agent log
        _qlog("list_quizzes_error", {"error": str(e)}, "H2")
        # #endregion
        raise


@router.get("/{quiz_id}")
def get_quiz(
    quiz_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get quiz detail with full questions."""
    quiz = db.query(GeneratedQuiz).options(joinedload(GeneratedQuiz.user)).filter(GeneratedQuiz.id == quiz_id).first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Không tìm thấy quiz")
    _require_course_access(db, current_user, quiz.course_id)

    result = _quiz_to_response(quiz)
    result["questions"] = quiz.questions or []
    return result


@router.post("/")
def create_quiz(
    payload: QuizCreateManual,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a quiz manually. Teacher or Admin only."""
    if current_user.role == UserRole.student:
        raise HTTPException(status_code=403, detail="Chỉ giáo viên hoặc admin mới có thể tạo quiz")
    _require_course_access(db, current_user, payload.course_id)

    questions = [q.model_dump() for q in payload.questions]
    quiz = GeneratedQuiz(
        course_id=payload.course_id,
        user_id=current_user.id,
        title=payload.title,
        questions=questions,
        source_chunks=[],
        is_ai_generated=0,
    )
    db.add(quiz)
    db.commit()
    db.refresh(quiz)
    logger.info(f"Manual quiz created: id={quiz.id}, by user={current_user.id}")
    _notify_students_new_quiz(db, quiz)
    return _quiz_to_response(quiz)


@router.post("/ai-generate")
async def create_quiz_ai(
    payload: QuizCreateAI,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Create a quiz using AI from course documents.
    Teacher or Admin only.
    """
    if current_user.role == UserRole.student:
        raise HTTPException(status_code=403, detail="Chỉ giáo viên hoặc admin mới có thể tạo quiz bằng AI")
    _require_course_access(db, current_user, payload.course_id)

    try:
        from app.services.tools.quiz_generator import QuizGenerator
        generator = QuizGenerator(db)
        result = await generator.generate(
            query=payload.topic or payload.title,
            course_id=payload.course_id,
            user_id=current_user.id,
            num_questions=payload.num_questions,
        )

        quiz_id = result["metadata"].get("quiz_id")
        if not quiz_id:
            raise HTTPException(status_code=500, detail=result.get("response", "Tạo quiz thất bại"))

        quiz = db.query(GeneratedQuiz).filter(GeneratedQuiz.id == quiz_id).first()
        if not quiz:
            raise HTTPException(status_code=500, detail="Không tìm thấy quiz vừa tạo")

        # Update title if user specified one
        if payload.title:
            quiz.title = payload.title
            db.commit()

        _notify_students_new_quiz(db, quiz)
        resp = _quiz_to_response(quiz)
        resp["questions"] = quiz.questions or []
        return resp

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"AI quiz generation error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi tạo quiz AI: {str(e)[:200]}")


@router.delete("/{quiz_id}", status_code=204)
def delete_quiz(
    quiz_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a quiz. Owner (teacher who created it) or Admin."""
    quiz = db.query(GeneratedQuiz).filter(GeneratedQuiz.id == quiz_id).first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Không tìm thấy quiz")

    is_owner = quiz.user_id == current_user.id
    is_admin = current_user.role == UserRole.admin
    if not (is_owner or is_admin):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xóa quiz này")

    db.delete(quiz)
    db.commit()
    logger.info(f"Quiz deleted: id={quiz_id}")


@router.post("/{quiz_id}/attempts")
def submit_attempt(
    quiz_id: int,
    payload: AttemptSubmit,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Submit a quiz attempt (student completes quiz)."""
    quiz = db.query(GeneratedQuiz).filter(GeneratedQuiz.id == quiz_id).first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Không tìm thấy quiz")
    _require_course_access(db, current_user, quiz.course_id)

    questions = quiz.questions or []
    total = len(questions)
    if total == 0:
        raise HTTPException(status_code=400, detail="Quiz không có câu hỏi")

    # Calculate score
    score = 0
    for idx, q in enumerate(questions):
        chosen = payload.answers.get(str(idx))
        if chosen and chosen.upper() == q.get("correct_answer", "").upper():
            score += 1

    started = None
    if payload.started_at:
        try:
            started = datetime.fromisoformat(payload.started_at)
        except ValueError:
            pass

    attempt = QuizAttempt(
        quiz_id=quiz_id,
        user_id=current_user.id,
        answers=payload.answers,
        score=score,
        total_questions=total,
        time_spent_seconds=payload.time_spent_seconds,
        started_at=started,
        completed_at=datetime.utcnow(),
    )
    db.add(attempt)
    db.commit()
    db.refresh(attempt)
    logger.info(f"Quiz attempt saved: attempt_id={attempt.id}, quiz={quiz_id}, user={current_user.id}, score={score}/{total}")

    return {
        **_attempt_to_response(attempt),
        "score_pct": round(score / total * 100, 1) if total > 0 else 0,
        "correct_indices": [
            idx for idx, q in enumerate(questions)
            if payload.answers.get(str(idx), "").upper() == q.get("correct_answer", "").upper()
        ],
    }


@router.get("/{quiz_id}/attempts/export")
def export_quiz_attempts_excel(
    quiz_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Xuất Excel kết quả tất cả lần làm của một quiz (giáo viên / admin)."""
    if current_user.role == UserRole.student:
        raise HTTPException(status_code=403, detail="Chỉ giáo viên hoặc admin mới có thể xem kết quả")

    quiz = (
        db.query(GeneratedQuiz)
        .options(joinedload(GeneratedQuiz.course))
        .filter(GeneratedQuiz.id == quiz_id)
        .first()
    )
    if not quiz:
        raise HTTPException(status_code=404, detail="Không tìm thấy quiz")

    _require_course_access(db, current_user, quiz.course_id)

    attempts = (
        db.query(QuizAttempt)
        .options(joinedload(QuizAttempt.user))
        .filter(QuizAttempt.quiz_id == quiz_id)
        .order_by(QuizAttempt.created_at.desc())
        .all()
    )

    course_name = quiz.course.name if quiz.course else f"Môn #{quiz.course_id}"
    content = _build_quiz_attempts_workbook(quiz.title, course_name, attempts)
    fname = f"quiz_{_safe_filename_segment(quiz.title, 50)}_{datetime.utcnow().strftime('%Y%m%d')}"
    return _excel_response(content, fname)


@router.get("/{quiz_id}/attempts")
def get_quiz_attempts(
    quiz_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get all attempts for a quiz. Teacher/Admin only."""
    quiz = db.query(GeneratedQuiz).filter(GeneratedQuiz.id == quiz_id).first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Không tìm thấy quiz")

    if current_user.role == UserRole.student:
        raise HTTPException(status_code=403, detail="Chỉ giáo viên hoặc admin mới có thể xem kết quả")

    _require_course_access(db, current_user, quiz.course_id)

    attempts = (
        db.query(QuizAttempt)
        .options(joinedload(QuizAttempt.user))
        .filter(QuizAttempt.quiz_id == quiz_id)
        .order_by(QuizAttempt.created_at.desc())
        .all()
    )
    return {"items": [_attempt_to_response(a) for a in attempts], "total": len(attempts)}
